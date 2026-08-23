"""County childcare prices, from the DOL Women's Bureau National Database of Childcare Prices.

Fills `family_childcare_cost_infant` and `family_childcare_cost_toddler`, the two indicators
that made `family_childcare` a domain where weight evaporated silently.

Prices are published **weekly** for full-time care and converted to annual here, because a
household compares childcare against a salary rather than against a week. Median centre-based
rates are used: family childcare homes are cheaper but far less uniformly available, so the
centre rate is the one a couple can actually plan around.

The database runs 2008-2018 and stops there. That is old enough to matter — childcare prices
rose sharply afterwards — so the vintage is recorded and the indicator should be read as
relative standing between counties rather than as a budget figure.
"""

from __future__ import annotations

from pathlib import Path

import polars as pl

from wlm.geo import is_in_scope, norm_fips
from wlm.ingest.base import emit

SOURCE_ID = "dol_childcare"
URL = "https://www.dol.gov/sites/dolgov/files/WB/media/nationaldatabaseofchildcareprices.xlsx"
LATEST_YEAR = 2018
WEEKS_PER_YEAR = 52

COLUMNS = {"MCInfant": "family_childcare_cost_infant",
           "MCToddler": "family_childcare_cost_toddler"}


def ingest(path: Path, *, year: int = LATEST_YEAR) -> tuple[pl.DataFrame, dict]:
    import openpyxl

    workbook = openpyxl.load_workbook(Path(path), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    index = {name: position for position, name in enumerate(header)}

    missing = [c for c in (*COLUMNS, "County_FIPS_Code", "StudyYear") if c not in index]
    if missing:
        workbook.close()
        raise ValueError(f"{SOURCE_ID}: workbook is missing {', '.join(missing)}")

    records: list[dict] = []
    counties: set[str] = set()
    for row in rows:
        if row[index["StudyYear"]] != year:
            continue
        raw = row[index["County_FIPS_Code"]]
        if raw in (None, ""):
            continue
        geoid = norm_fips(str(int(raw)), 5)
        if not is_in_scope(geoid):
            continue
        for column, indicator in COLUMNS.items():
            weekly = row[index[column]]
            if not isinstance(weekly, (int, float)) or weekly <= 0:
                continue
            records.append({"geo_level": "county", "geo_id": geoid,
                            "indicator_id": indicator,
                            "value": float(weekly) * WEEKS_PER_YEAR})
            counties.add(geoid)
    workbook.close()

    return emit(records, source_file=Path(path).name, vintage=str(year)), {
        "counties": len(counties),
        "year": year,
    }
